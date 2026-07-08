"""
Model cleaner — cleans imported CAD models by removing duplicate / irrelevant parts.

Algorithm:
  Step 0: Load STP + read XLSX J-column values
  Step 1: Match parts by J-column substring in part name → keep_step1
  Step 2: Interference check (AABB + boolean intersection, <3% volume) → keep_step2
  Step 3: Shape+position deduplication (>99% similarity) among kept parts
  Step 4: Output clean assembly (glb + assembly.json) + optional STEP export
"""

import os
import time
import logging

import numpy as np
from OCC.Core.Bnd import Bnd_Box
from OCC.Core.BRepBndLib import brepbndlib
from OCC.Core.BRepAlgoAPI import BRepAlgoAPI_Common
from OCC.Core.GProp import GProp_GProps
from OCC.Core.BRepGProp import brepgprop
from OCC.Core.TopoDS import TopoDS_Compound, TopoDS_Builder

from pipeline._occ_lock import OCC_BREP_LOCK

logger = logging.getLogger(__name__)


AABB_PADDING = 0.5
INTERFERENCE_THRESHOLD = 0.03


def _compute_aabb(shape):
    bbox = Bnd_Box()
    brepbndlib.Add(shape, bbox)
    xmin, ymin, zmin, xmax, ymax, zmax = bbox.Get()
    return np.array([xmin, ymin, zmin]), np.array([xmax, ymax, zmax])


def _aabbs_overlap(a_min, a_max, b_min, b_max):
    return bool(
        np.all(a_min <= b_max + AABB_PADDING) and
        np.all(a_max >= b_min - AABB_PADDING)
    )


def _compute_volume(shape):
    with OCC_BREP_LOCK:
        props = GProp_GProps()
        brepgprop.VolumeProperties(shape, props)
        return props.Mass()


def _compute_interference_volume(shape_a, shape_b):
    with OCC_BREP_LOCK:
        common = BRepAlgoAPI_Common(shape_a, shape_b)
        common.Build()
        if not common.IsDone():
            return 0.0
        result = common.Shape()
        if result.IsNull():
            return 0.0
        props = GProp_GProps()
        brepgprop.VolumeProperties(result, props)
        return props.Mass()


def _compute_centroid(shape):
    with OCC_BREP_LOCK:
        props = GProp_GProps()
        brepgprop.VolumeProperties(shape, props)
        if props.Mass() > 1e-12:
            c = props.CentreOfMass()
            return np.array([c.X(), c.Y(), c.Z()])
        brepgprop.SurfaceProperties(shape, props)
        c = props.CentreOfMass()
        return np.array([c.X(), c.Y(), c.Z()])


def _apply_transform_to_point(point_xyz, transform):
    if transform is None or len(transform) != 16:
        return point_xyz
    p = np.append(point_xyz, 1.0)
    m = np.array(transform, dtype=np.float64).reshape(4, 4, order='F')
    r = m @ p
    return r[:3]


DEDUP_POSITION_TOLERANCE_MM = 5.0
DEDUP_VOLUME_RATIO_MIN = 0.99
DEDUP_NEAR_MISS_AABB_DIM_MM = 5.0


def _shapes_equivalent(shape_a, shape_b, aabb_a, aabb_b,
                       centroid_a, centroid_b, vol_a, vol_b):
    if aabb_a is None or aabb_b is None:
        return False, "aabb_none"

    dim_a = aabb_a[1] - aabb_a[0]
    dim_b = aabb_b[1] - aabb_b[0]
    dim_diff = np.abs(dim_a - dim_b)
    max_dim_diff = float(np.max(dim_diff))
    if max_dim_diff > DEDUP_POSITION_TOLERANCE_MM:
        return False, "aabb_dim:%s" % ",".join("%.1f" % v for v in dim_diff)

    centroid_delta = -1.0
    if centroid_a is not None and centroid_b is not None:
        centroid_delta = float(np.linalg.norm(centroid_a - centroid_b))
        if centroid_delta > DEDUP_POSITION_TOLERANCE_MM:
            return False, "centroid:%.2f" % centroid_delta

    volume_ratio = 1.0
    if vol_a > 1e-12 and vol_b > 1e-12:
        volume_ratio = min(vol_a, vol_b) / max(vol_a, vol_b)
        if volume_ratio < DEDUP_VOLUME_RATIO_MIN:
            return False, "vol_ratio:%.4f" % volume_ratio

    return True, "ok"


def clean_model(stp_path, xlsx_path, output_dir, export_step=False, log_fn=None):
    _log = log_fn or print

    _log("=== 数模清洗 ===")
    _log("STP: {}".format(stp_path))
    _log("XLSX: {}".format(xlsx_path))
    _log("")

    # ── Step 0: Load ────────────────────────────────────────
    _log("[Step 0] 加载数据...")

    j_codes = _read_j_column(xlsx_path, _log)
    if not j_codes:
        _log("ERROR: XLSX J 列无有效数据")
        return 1
    _log("  J列共 {} 个代码: {}".format(len(j_codes), j_codes[:10]))

    from pipeline.stp_reader import read_stp_with_doc, verify_doc
    from pipeline.xcaf_utils import extract_assembly_tree, flatten_assembly_tree

    t0 = time.time()
    doc = read_stp_with_doc(stp_path)
    summary = verify_doc(doc, filepath=stp_path)
    _log("  读取 STP ({:.1f}s), Root shapes: {}".format(time.time() - t0, summary["root_count"]))
    if not summary["valid"]:
        _log("ERROR: 无有效形状")
        return 1

    roots = extract_assembly_tree(doc)
    parts, sub_assemblies = flatten_assembly_tree(roots)
    _log("  零件总数: {}".format(len(parts)))
    if len(parts) == 0:
        _log("ERROR: 无零件")
        return 1

    from pipeline.xcaf_utils import diagnose_assembly_tree
    diag = diagnose_assembly_tree(doc)
    _log("  Assembly节点: {}  |  Compound节点: {}  |  误判Compound: {}".format(
        diag["assembly_nodes"], diag["compound_nodes"],
        diag["misclassified_compounds"]))
    _log("  深度分布: {}".format(
        ", ".join("d{}={}".format(k, diag["depth_distribution"][k])
                  for k in sorted(diag["depth_distribution"].keys(), key=int))))
    if diag["misclassified_compounds"] > 0:
        _log("  WARNING: 有 {} 个节点被误判为Compound, 层级可能丢失".format(
            diag["misclassified_compounds"]))

    # ── Step 1: J-column matching ───────────────────────────
    _log("")
    _log("[Step 1] J列名称匹配...")

    keep_step1 = set()
    for i, part in enumerate(parts):
        name = part.get("name", "")
        for code in j_codes:
            if code in name:
                keep_step1.add(i)
                break

    _log("  匹配保留: {} / {}".format(len(keep_step1), len(parts)))

    part_names = [p.get("name", "?") for p in parts]
    if keep_step1:
        matched_names = [part_names[i] for i in sorted(keep_step1)[:10]]
        _log("  示例匹配: {}".format(matched_names))

    # ── Step 2: Interference check ──────────────────────────
    _log("")
    _log("[Step 2] 干涉检查 (阈值: {}% 自身体积)...".format(int(INTERFERENCE_THRESHOLD * 100)))

    keep_step2 = set()

    if not keep_step1:
        _log("  WARNING: Step 1 无匹配，跳过干涉检查")
    else:
        from pipeline import is_cancelled

        step1_indices = sorted(keep_step1)
        non_step1 = [i for i in range(len(parts)) if i not in keep_step1]

        self_volumes = {}
        for i in non_step1:
            try:
                self_volumes[i] = _compute_volume(parts[i]["shape"])
            except Exception:
                self_volumes[i] = 1.0

        aabbs = []
        for p in parts:
            try:
                aabbs.append(_compute_aabb(p["shape"]))
            except Exception:
                aabbs.append((np.zeros(3), np.zeros(3)))

        checked = 0
        kept_count = 0
        for i in non_step1:
            if is_cancelled():
                _log("  CANCELLED")
                return 1

            checked += 1
            if checked % 100 == 0 or checked == len(non_step1):
                _log("  AABB筛选: {}/{} (保留 {})".format(checked, len(non_step1), kept_count))

            a_min, a_max = aabbs[i]
            has_interference = False

            for j in step1_indices:
                b_min, b_max = aabbs[j]
                if not _aabbs_overlap(a_min, a_max, b_min, b_max):
                    continue

                try:
                    int_vol = _compute_interference_volume(parts[i]["shape"], parts[j]["shape"])
                except Exception:
                    continue

                self_vol = self_volumes.get(i, 1.0)
                if self_vol > 1e-12 and int_vol / self_vol >= INTERFERENCE_THRESHOLD:
                    has_interference = True
                    break

            if not has_interference:
                keep_step2.add(i)
                kept_count += 1

        _log("  干涉检查保留: {}".format(len(keep_step2)))

    # ── Step 3: Shape+position deduplication ────────────────
    _log("")
    _log("[Step 3] 形状+位置去重 (AABB + 质心 + 体积比例)...")

    all_kept = keep_step1 | keep_step2
    if not all_kept:
        _log("  无BOM匹配, 将对全部 {} 个零件进行形状+位置去重".format(len(parts)))
        all_kept = set(range(len(parts)))
    kept_list = sorted(all_kept)
    _log("  当前保留总数: {}".format(len(kept_list)))

    aabbs = []
    centroids = []
    volumes = []

    for idx in kept_list:
        shape = parts[idx]["shape"]
        xform = parts[idx].get("transform")
        try:
            aabbs.append(_compute_aabb(shape))
        except Exception:
            aabbs.append(None)
        try:
            local_centroid = _compute_centroid(shape)
            if local_centroid is not None and xform is not None:
                centroids.append(_apply_transform_to_point(local_centroid, xform))
            else:
                centroids.append(local_centroid)
        except Exception:
            centroids.append(None)
        try:
            volumes.append(_compute_volume(shape))
        except Exception:
            volumes.append(0.0)

    removed_by_dedup = set()
    n = len(kept_list)
    checked = 0
    fail_reasons = {}
    near_misses = []
    aabb_diff_buckets = [0] * 6

    for i_idx in range(n):
        if kept_list[i_idx] in removed_by_dedup:
            continue
        for j_idx in range(i_idx + 1, n):
            if kept_list[j_idx] in removed_by_dedup:
                continue

            checked += 1
            if checked % 500 == 0:
                _log("  去重检查: {}/{} 对 (已移除 {})".format(
                    checked, n * (n - 1) // 2, len(removed_by_dedup)))

            eq, reason = _shapes_equivalent(
                parts[kept_list[i_idx]]["shape"],
                parts[kept_list[j_idx]]["shape"],
                aabbs[i_idx], aabbs[j_idx],
                centroids[i_idx], centroids[j_idx],
                volumes[i_idx], volumes[j_idx],
            )

            if eq:
                removed_by_dedup.add(kept_list[j_idx])
            else:
                kind = reason.split(":")[0]
                fail_reasons[kind] = fail_reasons.get(kind, 0) + 1

                if centroids[i_idx] is not None and centroids[j_idx] is not None:
                    cd = float(np.linalg.norm(
                        centroids[i_idx] - centroids[j_idx]))
                else:
                    cd = -1.0

                if aabbs[i_idx] is not None and aabbs[j_idx] is not None:
                    dim_i = aabbs[i_idx][1] - aabbs[i_idx][0]
                    dim_j = aabbs[j_idx][1] - aabbs[j_idx][0]
                    dim_diff = np.abs(dim_i - dim_j)
                    max_ad = float(np.max(dim_diff))
                else:
                    max_ad = -1.0
                    dim_diff = None

                vr = 1.0
                if volumes[i_idx] > 1e-12 and volumes[j_idx] > 1e-12:
                    vr = min(volumes[i_idx], volumes[j_idx]) / max(
                        volumes[i_idx], volumes[j_idx])

                if cd >= 0 and max_ad >= 0 and max_ad <= DEDUP_NEAR_MISS_AABB_DIM_MM:
                    near_misses.append((
                        max_ad, cd, vr, reason,
                        part_names[kept_list[i_idx]],
                        part_names[kept_list[j_idx]],
                        dim_diff,
                    ))

                if kind == "aabb_dim" and max_ad >= 0:
                    if max_ad < 1:
                        aabb_diff_buckets[0] += 1
                    elif max_ad < 2:
                        aabb_diff_buckets[1] += 1
                    elif max_ad < 3:
                        aabb_diff_buckets[2] += 1
                    elif max_ad < 5:
                        aabb_diff_buckets[3] += 1
                    elif max_ad < 10:
                        aabb_diff_buckets[4] += 1
                    else:
                        aabb_diff_buckets[5] += 1

    _log("  检查对数: {} | 去重移除: {}".format(checked, len(removed_by_dedup)))
    if fail_reasons:
        _log("  失败原因分布: {}".format(
            ", ".join("{}={}".format(k, v)
                       for k, v in sorted(fail_reasons.items(),
                                          key=lambda x: -x[1]))))
    if any(aabb_diff_buckets):
        _log("  AABB尺寸差分布: <1mm={} 1-2mm={} 2-3mm={} 3-5mm={} 5-10mm={} >10mm={}".format(
            *aabb_diff_buckets))

    if near_misses:
        near_misses.sort(key=lambda x: (x[1], x[0]))
        _log("  最近似对 Top 20 (AABB差<{}mm):".format(
            DEDUP_NEAR_MISS_AABB_DIM_MM))
        for max_ad, cd, vr, reason, na, nb, dd in near_misses[:20]:
            dd_str = "(%s)" % ",".join("%.1f" % v for v in dd) if dd is not None else ""
            _log("    aabb=%.2fmm cent=%.2fmm vol=%.2f%% %s [%s] %s vs %s" % (
                max_ad, cd, vr * 100, dd_str, reason, na, nb))

    _log("  去重移除: {} 个".format(len(removed_by_dedup)))

    final_kept = all_kept - removed_by_dedup
    final_removed = set(range(len(parts))) - final_kept

    _log("  去重移除: {} 个".format(len(removed_by_dedup)))
    _log("  最终保留: {} / {}".format(len(final_kept), len(parts)))

    # ── Step 4: Output ──────────────────────────────────────
    _log("")
    _log("[Step 4] 输出清洗结果...")

    parts_dir = os.path.join(output_dir, "parts")
    os.makedirs(parts_dir, exist_ok=True)

    kept_parts = [parts[i] for i in sorted(final_kept)]
    filtered_roots = _filter_hierarchy(roots, final_kept, parts)

    from pipeline.gltf_exporter import export_assembly_indexed
    from pipeline.assembly_json import build_assembly_json, write_assembly_json

    t0 = time.time()
    kept_parts = export_assembly_indexed(kept_parts, parts_dir, linear_deflection=1.0)
    _log("  {} glb 文件 ({:.1f}s)".format(len(kept_parts), time.time() - t0))

    assembly = build_assembly_json(kept_parts, [], stp_path, roots=filtered_roots)
    json_path = os.path.join(output_dir, "assembly.json")
    write_assembly_json(assembly, json_path)
    _log("  assembly.json ({:.1f} KB)".format(os.path.getsize(json_path) / 1024))

    # ── Report ──────────────────────────────────────────────
    report_lines = []
    report_lines.append("=" * 60)
    report_lines.append("数模清洗报告")
    report_lines.append("=" * 60)
    report_lines.append("源文件: {}".format(stp_path))
    report_lines.append("BOM文件: {}".format(xlsx_path))
    report_lines.append("总零件数: {}".format(len(parts)))
    report_lines.append("-" * 60)
    report_lines.append("Step 1 (J列匹配):  保留 {}".format(len(keep_step1)))
    report_lines.append("Step 2 (干涉检查):  保留 {}".format(len(keep_step2)))
    report_lines.append("Step 3 (形状去重):  移除 {}".format(len(removed_by_dedup)))
    report_lines.append("-" * 60)
    report_lines.append("最终保留: {}".format(len(final_kept)))
    report_lines.append("最终清除: {}".format(len(final_removed)))
    report_lines.append("-" * 60)
    report_lines.append("保留的零件:")
    for i in sorted(final_kept):
        label = "[Step1]" if i in keep_step1 else ("[Step2]" if i in keep_step2 else "[?]")
        report_lines.append("  {} {:4d} | {}".format(label, i, part_names[i]))
    report_lines.append("-" * 60)
    report_lines.append("清除的零件:")
    for i in sorted(final_removed)[:50]:
        reason = "去重" if i in removed_by_dedup else ("干涉" if i not in (keep_step1 | keep_step2) else "未匹配")
        report_lines.append("  [{:5s}] {:4d} | {}".format(reason, i, part_names[i]))
    if len(final_removed) > 50:
        report_lines.append("  ... 及另外 {} 个".format(len(final_removed) - 50))
    report_lines.append("=" * 60)

    report = "\n".join(report_lines)
    report_path = os.path.join(output_dir, "clean_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)
    _log(report)

    # ── Optional STEP export ────────────────────────────────
    if export_step:
        _log("")
        _log("[Step 4b] 导出清洗后 STEP...")
        _export_clean_step(kept_parts, stp_path, output_dir, _log)

    _log("")
    _log("数模清洗完成. 输出: {}".format(output_dir))
    return 0


def _read_j_column(xlsx_path, log_fn=None):
    _log = log_fn or print
    try:
        from openpyxl import load_workbook
    except ImportError:
        _log("ERROR: openpyxl required")
        return []

    codes = []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if len(row) < 10:
                continue
            col_j = row[9].value
            if col_j is None:
                continue
            code = str(col_j).strip()
            if code:
                codes.append(code)
        wb.close()
    except Exception as e:
        _log("ERROR: 读取 XLSX 失败: {}".format(e))
        return []
    return codes


def _filter_hierarchy(roots, kept_part_indices, parts):
    kept_names = set()
    for i in kept_part_indices:
        kept_names.add(parts[i].get("name", "").replace(" ", "_"))

    def _filter_node(node):
        name = node.get("name", "").replace(" ", "_")
        children = node.get("children", [])
        filtered_children = []
        for child in children:
            filtered = _filter_node(child)
            if filtered is not None:
                filtered_children.append(filtered)

        is_in_kept = name in kept_names
        has_kept_descendant = len(filtered_children) > 0

        if is_in_kept or has_kept_descendant:
            result = dict(node)
            result["children"] = filtered_children
            if not is_in_kept and has_kept_descendant:
                part_ids = []
                for c in filtered_children:
                    part_ids.extend(c.get("partIds", []))
                result["partIds"] = part_ids
            return result
        return None

    result = []
    for root in roots:
        filtered = _filter_node(root)
        if filtered is not None:
            result.append(filtered)
    return result


def _export_clean_step(kept_parts, source_stp_path, output_dir, log_fn=None):
    _log = log_fn or print
    try:
        from OCC.Core.STEPControl import STEPControl_Writer
        from OCC.Core.Interface import Interface_Static_SetCVal
    except ImportError as e:
        _log("WARNING: STEP export not available: {}".format(e))
        return

    compound = TopoDS_Compound()
    builder = TopoDS_Builder()
    builder.MakeCompound(compound)

    solid_count = 0
    for part in kept_parts:
        shape = part.get("shape")
        if shape is None:
            continue
        from OCC.Core.TopAbs import TopAbs_SOLID
        from OCC.Core.TopExp import TopExp_Explorer
        exp = TopExp_Explorer(shape, TopAbs_SOLID)
        while exp.More():
            builder.Add(compound, exp.Current())
            solid_count += 1
            exp.Next()

    if solid_count == 0:
        _log("WARNING: no solids to export")
        return

    try:
        Interface_Static_SetCVal("write.step.schema", "AP214")
    except Exception:
        pass

    base_name = os.path.splitext(os.path.basename(source_stp_path))[0]
    step_path = os.path.join(output_dir, base_name + "_cleaned.stp")

    with OCC_BREP_LOCK:
        writer = STEPControl_Writer()
        writer.Transfer(compound, 0)
        status = writer.Write(step_path)

    if status == 1:
        size_kb = os.path.getsize(step_path) / 1024
        _log("  STEP 已导出: {} ({:.1f} KB)".format(step_path, size_kb))
    else:
        _log("  WARNING: STEP 导出失败 (status={})".format(status))
