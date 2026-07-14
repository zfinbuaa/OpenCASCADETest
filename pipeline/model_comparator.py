"""
Model comparator — compares two CAD assemblies at assembly-level.

Algorithm:
  1. Read .xlsx (Sheet3, col A = code_a, col B = code_b)
  2. Look up STP files by code in models_dir
  3. Load each STP → mesh all parts (world-space) → compute total volume
  4. Auto-align by centroid
  5. Step 1: Volume ratio quick check
  6. Step 2: Adaptive mesh sampling + bidirectional distance match
  7. Classify: 一致 / 细微差异 / 明显不一致
  8. Output compare_report.txt + compare_report.json
"""

import os
import re
import json
import time
import logging

import numpy as np
from OCC.Core.TopoDS import TopoDS_Compound, TopoDS_Builder
from OCC.Core.BRepBuilderAPI import BRepBuilderAPI_Transform
from OCC.Core.gp import gp_Trsf, gp_Vec
from OCC.Core.GProp import GProp_GProps
from OCC.Core.BRepGProp import brepgprop

from pipeline._occ_lock import OCC_BREP_LOCK

logger = logging.getLogger(__name__)

_SAFE_CODE_RE = re.compile(r'^[A-Za-z0-9_.\-]+$')
_COLUMN_HEADER_PATTERNS = [
    re.compile(r'^(code|部件代号|编码|代号|编号|号码)', re.IGNORECASE),
    re.compile(r'^(model|模型)', re.IGNORECASE),
]

SIMILARITY_IDENTICAL = 0.99
SIMILARITY_MINOR_DIFF = 0.95
VOLUME_RATIO_COARSE = 0.80

MESH_DEFLECTION = 1.0
SURFACE_DENSITY = 0.25
SAMPLE_MIN = 1000
SAMPLE_MAX = 50000
DISTANCE_THRESHOLD_MM = 2.0

MESH_CHUNK_SIZE = 64


def read_compare_xlsx(xlsx_path, models_dir=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl is required. Install: pip install openpyxl")
        return []

    if models_dir is None:
        models_dir = os.path.dirname(os.path.abspath(xlsx_path))

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("failed to open xlsx: %s", e)
        return []

    pairs = []
    try:
        sheet_names = wb.sheetnames
        target_sheet = None
        for name in sheet_names:
            if name.lower().replace(' ', '') == 'sheet3' or name == '3' or name == 'Sheet3':
                target_sheet = name
                break
        if target_sheet is None:
            for idx, name in enumerate(sheet_names):
                if idx == 2:
                    target_sheet = name
                    break
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]
            logger.warning("Sheet3 not found, using first sheet: %s", target_sheet)

        ws = wb[target_sheet]
        start_row = 1
        row_count = 0
        for row in ws.iter_rows(min_row=1):
            row_count += 1
            if row_count > 2000:
                break
            val_a = None
            val_b = None
            if len(row) >= 1:
                val_a = row[0].value
            if len(row) >= 2:
                val_b = row[1].value

            if val_a is None and val_b is None:
                continue

            code_a = str(val_a).strip() if val_a is not None else ""
            code_b = str(val_b).strip() if val_b is not None else ""

            if not code_a or not code_b:
                continue

            if start_row == 1 and _looks_like_header(code_a) and _looks_like_header(code_b):
                start_row = 2
                continue

            if not _SAFE_CODE_RE.match(code_a):
                logger.warning("skip compare row with unsafe code_a: %r", code_a)
                continue
            if not _SAFE_CODE_RE.match(code_b):
                logger.warning("skip compare row with unsafe code_b: %r", code_b)
                continue

            stp_a = _find_stp(code_a, models_dir)
            stp_b = _find_stp(code_b, models_dir)

            pairs.append((code_a, stp_a, code_b, stp_b, len(pairs) + 1))
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return pairs


def _looks_like_header(val):
    s = str(val).strip().lower()
    for pat in _COLUMN_HEADER_PATTERNS:
        if pat.match(s):
            return True
    common = {'code', 'a', 'b', 'model', 'id', 'name', '零件', '部件'}
    if s in common:
        return True
    return False


_SEPARATORS = frozenset(['-', '_', '.', ' '])


def _find_stp(code, models_dir):
    for ext in (".stp", ".STP", ".step", ".STEP", ".Step"):
        candidate = os.path.join(models_dir, code + ext)
        if os.path.exists(candidate):
            return candidate

    code_lower = code.lower()
    best = None

    try:
        for entry in os.listdir(models_dir):
            base, ext = os.path.splitext(entry)
            if ext.lower() not in ('.stp', '.step'):
                continue
            base_lower = base.lower()
            if base_lower == code_lower:
                return os.path.join(models_dir, entry)
            if not base_lower.startswith(code_lower):
                continue
            if len(base_lower) == len(code_lower):
                continue
            sep_char = base_lower[len(code_lower)]
            if sep_char not in _SEPARATORS:
                continue
            if best is None or len(base) < len(best[0]):
                best = (base, entry)

    except OSError:
        pass

    if best is not None:
        logger.info("prefix match '%s' -> '%s'", code, best[0])
        return os.path.join(models_dir, best[1])

    return None


def _load_assembly(stp_path, log_fn=None):
    from pipeline.stp_reader import read_stp_with_doc, verify_doc
    from pipeline.xcaf_utils import extract_assembly_tree, flatten_assembly_tree

    _log = log_fn or (lambda m: None)

    doc = read_stp_with_doc(stp_path)
    summary = verify_doc(doc, filepath=stp_path)
    if not summary["valid"]:
        raise RuntimeError("No valid shapes in: " + stp_path)

    roots = extract_assembly_tree(doc)
    parts, sub_assemblies = flatten_assembly_tree(roots)

    if not parts:
        raise RuntimeError("No parts found in: " + stp_path)

    depths = set()
    for sa in sub_assemblies:
        d = sa.get("depth", 0)
        depths.add(d)

    stats = {
        "part_count": len(parts),
        "sub_assembly_count": len(sub_assemblies),
        "max_depth": max(depths) if depths else 0,
    }

    return parts, stats


def _mesh_part(part, linear_deflection=1.0):
    from pipeline.mesher import brep_to_mesh
    from pipeline.gltf_exporter import _apply_transform

    shape = part.get("shape")
    if shape is None:
        return None, None, None

    try:
        verts, tris, normals = brep_to_mesh(shape, linear_deflection=linear_deflection)
        if len(verts) < 9 or len(tris) < 1:
            return None, None, None
        v = np.array(verts, dtype=np.float64).reshape(-1, 3)
        t_arr = np.array(tris, dtype=np.int32)
        transform = part.get("transform")
        if transform and len(transform) == 16:
            v = _apply_transform(v, transform)
        return v, t_arr, shape
    except Exception:
        return None, None, None


def _compute_volume(shape):
    try:
        with OCC_BREP_LOCK:
            props = GProp_GProps()
            brepgprop.VolumeProperties(shape, props)
            return props.Mass()
    except Exception:
        return 0.0


def _build_full_mesh(parts, linear_deflection=1.0):
    from pipeline.mesher import brep_to_mesh
    from pipeline.gltf_exporter import _apply_transform

    all_positions = []
    all_indices = []
    vertex_offset = 0

    for p in parts:
        shape = p.get("shape")
        if shape is None:
            continue
        transform = p.get("transform")

        try:
            verts, tris, _ = brep_to_mesh(shape, linear_deflection=linear_deflection)
            if len(verts) < 9 or len(tris) < 1:
                continue
            v = np.array(verts, dtype=np.float64).reshape(-1, 3)
            t = np.array(tris, dtype=np.int32).reshape(-1, 3)

            if transform and len(transform) == 16:
                v = _apply_transform(v, transform)

            all_positions.append(v)
            all_indices.append(t + vertex_offset)
            vertex_offset += len(v)
        except Exception:
            continue

    if not all_positions:
        return np.empty((0, 3), dtype=np.float32), np.empty((0, 3), dtype=np.int32)

    positions = np.vstack(all_positions).astype(np.float32)
    indices = np.vstack(all_indices).astype(np.int32)
    return positions, indices


def _build_grid_index(points, cell_size):
    if len(points) == 0:
        return {}, cell_size

    min_v = points.min(axis=0)
    grid = {}
    for i, pt in enumerate(points):
        cx = int(np.floor((pt[0] - min_v[0]) / cell_size))
        cy = int(np.floor((pt[1] - min_v[1]) / cell_size))
        cz = int(np.floor((pt[2] - min_v[2]) / cell_size))
        key = (cx, cy, cz)
        if key not in grid:
            grid[key] = []
        grid[key].append(i)

    return grid, min_v


def _grid_nearest_distance(query_pts, target_pts, target_grid, cell_size, target_min):
    n = len(query_pts)
    if n == 0 or len(target_pts) == 0:
        return np.full(n, 1e18, dtype=np.float64)

    distances = np.full(n, 1e18, dtype=np.float64)
    search_radius_cells = max(1, int(np.ceil(DISTANCE_THRESHOLD_MM * 3 / cell_size)))

    for chunk_start in range(0, n, MESH_CHUNK_SIZE):
        chunk_end = min(chunk_start + MESH_CHUNK_SIZE, n)
        chunk = query_pts[chunk_start:chunk_end]
        chunk_dists = np.full(chunk_end - chunk_start, 1e18, dtype=np.float64)

        for j, pt in enumerate(chunk):
            cx = int(np.floor((pt[0] - target_min[0]) / cell_size))
            cy = int(np.floor((pt[1] - target_min[1]) / cell_size))
            cz = int(np.floor((pt[2] - target_min[2]) / cell_size))

            best = 1e18
            for dx in range(-search_radius_cells, search_radius_cells + 1):
                for dy in range(-search_radius_cells, search_radius_cells + 1):
                    for dz in range(-search_radius_cells, search_radius_cells + 1):
                        key = (cx + dx, cy + dy, cz + dz)
                        indices = target_grid.get(key)
                        if indices is None:
                            continue
                        diff = target_pts[indices] - pt
                        sq = np.dot(diff, diff.T) if len(diff.shape) == 1 and len(indices) == 1 else np.sum(diff * diff, axis=1)
                        min_sq = np.min(sq)
                        if min_sq < best:
                            best = min_sq

            chunk_dists[j] = np.sqrt(best)

        distances[chunk_start:chunk_end] = chunk_dists

    return distances


def _mesh_similarity(verts_a, verts_b):
    if len(verts_a) < 3 or len(verts_b) < 3:
        return 0.0, np.array([]), np.array([])

    merged = np.vstack([verts_a, verts_b])
    merged_min = merged.min(axis=0)
    merged_max = merged.max(axis=0)
    extents = merged_max - merged_min
    typical_extent = np.median(extents[extents > 0]) if np.any(extents > 0) else 100.0
    cell_size = max(5.0, typical_extent / 50.0)

    grid_b, min_b = _build_grid_index(verts_b, cell_size)
    dists_a = _grid_nearest_distance(verts_a, verts_b, grid_b, cell_size, min_b)
    match_a = np.mean(dists_a < DISTANCE_THRESHOLD_MM)

    grid_a, min_a = _build_grid_index(verts_a, cell_size)
    dists_b = _grid_nearest_distance(verts_b, verts_a, grid_a, cell_size, min_a)
    match_b = np.mean(dists_b < DISTANCE_THRESHOLD_MM)

    return float(min(match_a, match_b)), dists_a, dists_b


def _classify_result(similarity):
    if similarity >= SIMILARITY_IDENTICAL:
        return "一致"
    if similarity >= SIMILARITY_MINOR_DIFF:
        return "细微差异"
    return "明显不一致"


def compare_assemblies(stp_path_a, stp_path_b, code_a, code_b, output_dir, log_fn=None, diff_glb=False):
    _log = log_fn or print

    _log("")
    _log("-" * 60)
    _log("对比: {} vs {}".format(code_a, code_b))
    _log("  A: {}".format(stp_path_a or "(未找到)"))
    _log("  B: {}".format(stp_path_b or "(未找到)"))

    if stp_path_a is None or stp_path_b is None:
        missing = []
        if stp_path_a is None:
            missing.append(code_a)
        if stp_path_b is None:
            missing.append(code_b)
        _log("  ERROR: STP文件未找到: {}".format(", ".join(missing)))
        return {
            "code_a": code_a, "code_b": code_b,
            "stp_a": stp_path_a, "stp_b": stp_path_b,
            "error": "stp_not_found",
            "missing": missing,
        }

    t0 = time.time()

    try:
        _log("  [1/5] 加载模型 A...")
        parts_a, stats_a = _load_assembly(stp_path_a, log_fn=_log)
        _log("    零件数: {}, 子装配: {}, 最大深度: {}".format(
            stats_a["part_count"], stats_a["sub_assembly_count"], stats_a["max_depth"]))

        _log("  [2/5] 加载模型 B...")
        parts_b, stats_b = _load_assembly(stp_path_b, log_fn=_log)
        _log("    零件数: {}, 子装配: {}, 最大深度: {}".format(
            stats_b["part_count"], stats_b["sub_assembly_count"], stats_b["max_depth"]))
    except Exception as e:
        _log("  ERROR: 加载失败: {}".format(e))
        return {
            "code_a": code_a, "code_b": code_b,
            "stp_a": stp_path_a, "stp_b": stp_path_b,
            "error": "load_failed",
            "message": str(e),
        }

    _log("  [3/5] 网格化 + 体积计算...")
    t_mesh = time.time()
    mesh_a, vol_a = _mesh_all_parts(parts_a, _log, "A")
    mesh_b, vol_b = _mesh_all_parts(parts_b, _log, "B")
    _log("    体积A: {:.1f} mm³  体积B: {:.1f} mm³ (网格化 {:.1f}s)".format(
        vol_a, vol_b, time.time() - t_mesh))

    centroid_a = np.mean(mesh_a, axis=0) if len(mesh_a) > 0 else np.zeros(3)
    centroid_b = np.mean(mesh_b, axis=0) if len(mesh_b) > 0 else np.zeros(3)
    offset = centroid_a - centroid_b
    if len(mesh_b) > 0:
        mesh_b = mesh_b + offset

    _log("  [4/5] 体积比粗筛...")
    max_vol = max(vol_a, vol_b)
    if max_vol < 1e-12:
        similarity_vol = 0.0
    else:
        similarity_vol = min(vol_a, vol_b) / max_vol

    _log("    体积比: {:.2%}".format(similarity_vol))

    if similarity_vol < VOLUME_RATIO_COARSE:
        similarity = 0.0
        classification = "明显不一致"
        mesh_sim = 0.0
        sample_count = 0
        dists_a = np.array([])
        dists_b = np.array([])
        _log("    体积差异 > {:.0%}, 跳过网格采样".format(1 - VOLUME_RATIO_COARSE))
    else:
        _log("  [5/5] 自适应网格采样 + 距离场对比...")
        t_samp = time.time()
        similarity, mesh_sim, sample_count, dists_a, dists_b = _adaptive_mesh_compare(
            mesh_a, mesh_b, vol_a, vol_b)
        classification = _classify_result(similarity)
        _log("    采样数: {}  双向匹配率: {:.2%}  判定相似度: {:.2%} ({:.1f}s)".format(
            sample_count, mesh_sim, similarity, time.time() - t_samp))

    # Write diff GLB if enabled
    diff_glb_a_path = None
    diff_glb_b_path = None
    if diff_glb:
        _log("    导出差异模型...")
        diff_glb_a_path, diff_glb_b_path = _export_diff_models(
            parts_a, parts_b, mesh_a, mesh_b, dists_a, dists_b,
            offset, output_dir, code_a, code_b, _log)

    metrics = {
        "volume_a": round(vol_a, 2),
        "volume_b": round(vol_b, 2),
        "volume_ratio": round(similarity_vol, 6),
        "mesh_similarity": round(mesh_sim, 6),
        "similarity": round(similarity, 6),
        "sample_count": sample_count,
        "align_offset": [round(x, 3) for x in offset.tolist()],
        "diff_glb_a": diff_glb_a_path,
        "diff_glb_b": diff_glb_b_path,
    }

    _log("  判定: {} (相似度 {:.2%})".format(classification, similarity))

    result = {
        "code_a": code_a,
        "code_b": code_b,
        "stp_a": stp_path_a,
        "stp_b": stp_path_b,
        "classification": classification,
        "structural": {
            "part_count_a": stats_a["part_count"],
            "part_count_b": stats_b["part_count"],
            "max_depth_a": stats_a["max_depth"],
            "max_depth_b": stats_b["max_depth"],
        },
        "geometric": metrics,
        "elapsed_s": round(time.time() - t0, 1),
    }

    _save_comparison_report(result, output_dir)
    return result


def _export_diff_models(parts_a, parts_b, mesh_a, mesh_b, dists_a, dists_b,
                        offset, output_dir, code_a, code_b, _log):
    glb_a_path = None
    glb_b_path = None

    try:
        if len(dists_a) >= 3 and len(mesh_a) >= 3:
            _log("    生成差异模型 A...")
            positions_a, indices_a = _build_full_mesh(parts_a, linear_deflection=MESH_DEFLECTION)
            if len(positions_a) >= 3:
                vert_dists_a = _propagate_distances(mesh_a, dists_a, positions_a)
                name_a = _safe_filename(code_a) + "_diff"
                glb_a_path = os.path.join(output_dir, name_a + ".glb")
                _write_diff_glb(positions_a, indices_a, vert_dists_a, glb_a_path, name=name_a)
                _log("      → {}".format(os.path.basename(glb_a_path)))
    except Exception as e:
        _log("    WARNING: 差异模型 A 生成失败: {}".format(e))

    try:
        if len(dists_b) >= 3 and len(mesh_b) >= 3:
            _log("    生成差异模型 B...")
            positions_b, indices_b = _build_full_mesh(parts_b, linear_deflection=MESH_DEFLECTION)
            if len(positions_b) >= 3:
                vert_dists_b = _propagate_distances(mesh_b, dists_b, positions_b)
                name_b = _safe_filename(code_b) + "_diff"
                glb_b_path = os.path.join(output_dir, name_b + ".glb")
                _write_diff_glb(positions_b, indices_b, vert_dists_b, glb_b_path, name=name_b)
                _log("      → {}".format(os.path.basename(glb_b_path)))
    except Exception as e:
        _log("    WARNING: 差异模型 B 生成失败: {}".format(e))

    return glb_a_path, glb_b_path


def _propagate_distances(sample_verts, sample_dists, full_verts):
    if len(sample_verts) < 3 or len(full_verts) < 3:
        return np.full(len(full_verts), 0.0, dtype=np.float64)

    n_full = len(full_verts)
    out = np.zeros(n_full, dtype=np.float64)

    chunk = 4096
    for start in range(0, n_full, chunk):
        end = min(start + chunk, n_full)
        diff = sample_verts[np.newaxis, :, :] - full_verts[start:end, np.newaxis, :]
        sq = np.sum(diff * diff, axis=2)
        min_idx = np.argmin(sq, axis=1)
        out[start:end] = sample_dists[min_idx]

    return out


def _safe_filename(code):
    """Strip unsafe characters from a filename component."""
    return re.sub(r'[\\/:*?"<>|\s]+', '_', code).strip('_')[:120]


def _mesh_all_parts(parts, log_fn, label):
    total_vol = 0.0
    all_verts = []

    n = len(parts)
    report_interval = max(1, min(n // 5, 100))
    meshed = 0

    for i, p in enumerate(parts):
        v, t_arr, shape = _mesh_part(p, linear_deflection=MESH_DEFLECTION)
        if v is not None:
            all_verts.append(v)
            meshed += 1
        if shape is not None:
            vol = _compute_volume(shape)
            total_vol += abs(vol)

        if n > 10 and (i + 1) % report_interval == 0:
            log_fn("    网格化 {}/{}...".format(i + 1, n))

    if all_verts:
        merged = np.vstack(all_verts)
    else:
        merged = np.empty((0, 3), dtype=np.float64)

    return merged, total_vol


def _compute_surface_area(vertices):
    if len(vertices) < 3:
        return 0.0

    min_v = vertices.min(axis=0)
    max_v = vertices.max(axis=0)
    extents = max_v - min_v
    bbox_surface = 2.0 * (extents[0] * extents[1] + extents[1] * extents[2] + extents[0] * extents[2])

    convex_hull_extent = 0.0
    for axis in range(3):
        projected = np.delete(vertices, axis, axis=1)
        if len(projected) >= 3:
            hull_min = projected.min(axis=0)
            hull_max = projected.max(axis=0)
            hull_area = (hull_max[0] - hull_min[0]) * (hull_max[1] - hull_min[1])
            convex_hull_extent += hull_area
    convex_hull_extent *= 2.0

    return max(bbox_surface, convex_hull_extent)


def _adaptive_mesh_compare(mesh_a, mesh_b, vol_a, vol_b):
    if len(mesh_a) < 3 or len(mesh_b) < 3:
        return 0.0, 0.0, 0, np.array([]), np.array([])

    surf_a = _compute_surface_area(mesh_a)
    surf_b = _compute_surface_area(mesh_b)
    avg_surface = max(surf_a, surf_b, 1.0)

    n_samples = int(np.clip(avg_surface * SURFACE_DENSITY, SAMPLE_MIN, SAMPLE_MAX))

    if len(mesh_a) > n_samples:
        idx_a = np.random.choice(len(mesh_a), n_samples, replace=False)
        sampled_a = mesh_a[idx_a]
    else:
        sampled_a = mesh_a
        n_samples = len(sampled_a)

    if len(mesh_b) > n_samples:
        idx_b = np.random.choice(len(mesh_b), n_samples, replace=False)
        sampled_b = mesh_b[idx_b]
    else:
        sampled_b = mesh_b

    mesh_sim, dists_a, dists_b = _mesh_similarity(sampled_a, sampled_b)

    max_vol = max(vol_a, vol_b)
    if max_vol < 1e-12:
        similarity_vol = 0.0
    else:
        similarity_vol = min(vol_a, vol_b) / max_vol

    similarity = 0.4 * similarity_vol + 0.6 * mesh_sim

    return similarity, mesh_sim, n_samples, dists_a, dists_b


def _write_diff_glb(positions, indices, distances, output_path, name="diff"):
    """
    Write a color-coded GLB where vertex colors encode distance-to-other-model.

    Green (< 2mm) → Yellow (2-5mm) → Red (> 5mm)
    """
    import struct

    if len(positions) < 3 or len(indices) < 1:
        return None

    verts_f32 = positions.astype(np.float32)
    indices_u32 = indices.flatten().astype(np.uint32)

    colors = np.zeros(len(verts_f32), dtype=np.uint8)
    for i, d in enumerate(distances):
        idx = i * 3
        if d < 2.0:
            colors[idx] = 0
            colors[idx + 1] = 200
            colors[idx + 2] = 0
        elif d < 5.0:
            colors[idx] = 255
            colors[idx + 1] = 215
            colors[idx + 2] = 0
        else:
            colors[idx] = 220
            colors[idx + 1] = 0
            colors[idx + 2] = 0

    pos_bytes = verts_f32.tobytes()
    col_bytes = colors.tobytes()
    idx_bytes = indices_u32.tobytes()

    def _pad(n):
        return (n + 3) // 4 * 4

    pos_len = len(pos_bytes)
    col_len = len(col_bytes)
    idx_len = len(idx_bytes)
    pos_pad = _pad(pos_len)
    col_pad = _pad(col_len)
    idx_pad = _pad(idx_len)

    total_bin_len = pos_pad + col_pad + idx_pad

    bbox_min = verts_f32.min(axis=0).tolist()
    bbox_max = verts_f32.max(axis=0).tolist()

    gltf = {
        "asset": {"version": "2.0", "generator": "AutoModel"},
        "scene": 0,
        "scenes": [{"name": "Scene", "nodes": [0]}],
        "nodes": [{"mesh": 0, "name": name}],
        "meshes": [{
            "name": name,
            "primitives": [{
                "attributes": {"POSITION": 1, "COLOR_0": 2},
                "indices": 0,
                "material": 0,
            }],
        }],
        "accessors": [
            {
                "bufferView": 0, "componentType": 5125,
                "count": len(indices_u32), "type": "SCALAR",
                "max": [int(indices_u32.max())], "min": [0],
            },
            {
                "bufferView": 1, "componentType": 5126,
                "count": len(verts_f32), "type": "VEC3",
                "max": bbox_max, "min": bbox_min,
            },
            {
                "bufferView": 2, "componentType": 5121,
                "count": len(verts_f32), "type": "VEC3",
                "max": [255, 255, 255], "min": [0, 0, 0],
                "normalized": True,
            },
        ],
        "bufferViews": [
            {"buffer": 0, "byteOffset": pos_pad + col_pad, "byteLength": idx_len, "target": 34963},
            {"buffer": 0, "byteOffset": 0, "byteLength": pos_len, "target": 34962},
            {"buffer": 0, "byteOffset": pos_pad, "byteLength": col_len, "target": 34962},
        ],
        "buffers": [{"byteLength": total_bin_len}],
        "materials": [{
            "pbrMetallicRoughness": {
                "baseColorFactor": [1.0, 1.0, 1.0, 1.0],
                "metallicFactor": 0.0,
                "roughnessFactor": 0.5,
            },
            "name": "diff",
            "doubleSided": True,
        }],
    }

    gltf_json = json.dumps(gltf, ensure_ascii=False)
    gltf_json += " " * (_pad(len(gltf_json)) - len(gltf_json))

    header = struct.pack('<I', 0x46546C67)
    header += struct.pack('<I', 2)
    header += struct.pack('<I', 12 + 8 + len(gltf_json) + 8 + total_bin_len)

    chunk_header = struct.pack('<I', len(gltf_json)) + struct.pack('<I', 0x4E4F534A)
    bin_header = struct.pack('<I', total_bin_len) + struct.pack('<I', 0x004E4942)
    bin_padding = b'\x00' * (pos_pad - pos_len)
    col_padding = b'\x00' * (col_pad - col_len)
    idx_padding = b'\x00' * (idx_pad - idx_len)
    bin_data = pos_bytes + bin_padding + col_bytes + col_padding + idx_bytes + idx_padding

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    with open(output_path, 'wb') as f:
        f.write(header)
        f.write(chunk_header)
        f.write(gltf_json.encode('utf-8'))
        f.write(bin_header)
        f.write(bin_data)

    return output_path


def _save_comparison_report(result, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    json_path = os.path.join(output_dir, "compare_report.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    txt_path = os.path.join(output_dir, "compare_report.txt")
    lines = []
    lines.append("=" * 60)
    lines.append("数模对比报告")
    lines.append("=" * 60)
    lines.append("模型A: {} → {}".format(result.get("code_a", "?"), result.get("stp_a", "?")))
    lines.append("模型B: {} → {}".format(result.get("code_b", "?"), result.get("stp_b", "?")))
    lines.append("-" * 60)
    lines.append("[结构指标]")
    s = result.get("structural", {})
    lines.append("  零件数:  A={}  B={}  (差异: {})".format(
        s.get("part_count_a", 0), s.get("part_count_b", 0),
        s.get("part_count_b", 0) - s.get("part_count_a", 0)))
    lines.append("  最大层级深度:  A={}  B={}".format(
        s.get("max_depth_a", 0), s.get("max_depth_b", 0)))
    lines.append("-" * 60)
    lines.append("[几何指标]（质心对齐后）")
    g = result.get("geometric", {})
    lines.append("  A总体积:     {:.1f} mm³".format(g.get("volume_a", 0)))
    lines.append("  B总体积:     {:.1f} mm³".format(g.get("volume_b", 0)))
    vol_a = g.get("volume_a", 1)
    if vol_a > 0:
        vol_diff = abs(g.get("volume_a", 0) - g.get("volume_b", 0)) / vol_a
        lines.append("  体积差:      {:.2%}".format(vol_diff))
    lines.append("  体积比:      {:.2%}".format(g.get("volume_ratio", 0)))
    lines.append("  网格匹配率:  {:.2%}".format(g.get("mesh_similarity", 0)))
    lines.append("  采样点数:    {}".format(g.get("sample_count", 0)))
    offset = g.get("align_offset", [0, 0, 0])
    lines.append("  质心偏移(mm): X={:.1f} Y={:.1f} Z={:.1f}".format(offset[0], offset[1], offset[2]))
    lines.append("-" * 60)
    lines.append("[判定]")
    lines.append("  对比结果: {}".format(result.get("classification", "?")))
    lines.append("  综合相似度: {:.2%}".format(g.get("similarity", 0)))
    cls = result.get("classification", "")
    if cls == "一致":
        lines.append("  说明:     两个模型几何尺寸一致。")
    elif cls == "细微差异":
        lines.append("  说明:     总体一致，有少量体积差异（可能由部件刻印等引起）。")
    else:
        lines.append("  说明:     两个模型存在明显形态或尺寸差异。")
    lines.append("  耗时:      {:.1f}s".format(result.get("elapsed_s", 0)))
    lines.append("=" * 60)

    report = "\n".join(lines)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(report)
