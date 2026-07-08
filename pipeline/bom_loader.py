"""
BOM (Bill of Materials) loader — reads Excel .xlsx files.

Expected Excel format:
  Column H = part name (部件名称)
  Column J = part code (部件编号), matches .stp filename without extension

Assembly tree matching:
  Part names in STP files follow the pattern: {J列}-{版本}-{H列}
  e.g. ABC123-V2.0-A前保险杠

Returns a list of BOM entries suitable for the multi-file loading pipeline.
"""

import os
import sys
import re
import logging

logger = logging.getLogger(__name__)

_SAFE_CODE_RE = re.compile(r'^[A-Za-z0-9_.\-]+$')


def read_bom(xlsx_path, models_dir=None):
    """
    Read BOM from an Excel .xlsx file.

    Args:
        xlsx_path: path to .xlsx file.
        models_dir: directory containing .stp files named by J-column codes.
                    If None, uses the directory of the xlsx file.

    Returns:
        list[dict]: each entry has keys:
            code (str):        J-column value (used to find STP file)
            target_name (str): H-column value (used to match node in assembly tree)
            stp_path (str):    resolved path to .stp file (or None if missing)
            exists (bool):     whether the .stp file was found
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl is required. Install: pip install openpyxl")
        return []

    if models_dir is None:
        models_dir = os.path.dirname(os.path.abspath(xlsx_path))
    real_models = os.path.realpath(models_dir)

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("failed to open BOM xlsx: %s", e)
        return []

    entries = []
    try:
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if len(row) < 10:
                continue

            col_h = row[7].value   # Column H (0-indexed: 7)
            col_j = row[9].value   # Column J (0-indexed: 9)

            if col_h is None and col_j is None:
                continue

            code = str(col_j).strip() if col_j is not None else ""
            target_name = str(col_h).strip() if col_h is not None else ""

            if not code and not target_name:
                continue

            if not code:
                entries.append({
                    "code": "",
                    "target_name": target_name,
                    "stp_path": None,
                    "exists": False,
                })
                continue

            # Reject codes that could escape models_dir or contain unsafe chars
            if not _SAFE_CODE_RE.match(code):
                logger.warning("skip BOM row with unsafe code: %r", code)
                continue

            stp_path = None
            for ext in (".stp", ".STP", ".step", ".STEP", ".Step"):
                candidate = os.path.join(models_dir, code + ext)
                if os.path.exists(candidate):
                    stp_path = candidate
                    break

            if stp_path is None:
                try:
                    code_lower = code.lower()
                    best = None
                    for entry in os.listdir(models_dir):
                        base, ext = os.path.splitext(entry)
                        if ext.lower() not in ('.stp', '.step'):
                            continue
                        base_lower = base.lower()
                        if base_lower == code_lower:
                            stp_path = os.path.join(models_dir, entry)
                            break
                        if not base_lower.startswith(code_lower):
                            continue
                        if len(base_lower) == len(code_lower):
                            continue
                        sep_char = base_lower[len(code_lower)]
                        if sep_char not in ('-', '_', '.', ' '):
                            continue
                        if best is None or len(base) < len(best[0]):
                            best = (base, entry)

                    if stp_path is None and best is not None:
                        logger.info("prefix match '%s' -> '%s'", code, best[0])
                        stp_path = os.path.join(models_dir, best[1])
                except OSError:
                    pass

            if stp_path is None:
                stp_path = os.path.join(models_dir, code + ".stp")

            # Defense-in-depth: ensure resolved path stays under models_dir
            try:
                real_stp = os.path.realpath(stp_path)
                if os.path.commonpath([real_models, real_stp]) != real_models:
                    logger.warning("skip BOM code escaping models_dir: %r", code)
                    continue
            except (ValueError, OSError):
                continue

            exists = os.path.exists(stp_path)

            if not target_name and code:
                target_name = code

            entries.append({
                "code": code,
                "target_name": target_name,
                "stp_path": stp_path,
                "exists": exists,
            })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return entries


def validate_bom_entries(entries, models_dir=None):
    """
    Validate BOM entries and return a report.

    Args:
        entries: list from read_bom().
        models_dir: optional directory path to list available STP files for hints.

    Returns:
        tuple: (valid_entries, missing_files, report_lines)
            valid_entries: entries with existing STP files.
            missing_files: list of entries for missing files.
            report_lines: list of string messages.
    """
    valid = []
    missing = []
    lines = []

    lines.append("BOM entries: {} total".format(len(entries)))

    for e in entries:
        if e["exists"]:
            valid.append(e)
        else:
            line = "  MISSING: '{}' ({})".format(
                e.get("target_name", ""), e.get("code", ""))
            if e.get("stp_path"):
                line += " -> {}".format(e["stp_path"])
            lines.append(line)
            missing.append(e)

    lines.append("  Valid: {} / Missing: {}".format(len(valid), len(missing)))

    if missing and models_dir and os.path.isdir(models_dir):
        try:
            avail = sorted([f for f in os.listdir(models_dir)
                            if f.lower().endswith(('.stp', '.step'))])[:20]
            if avail:
                lines.append("  Available STP/STEP files in models_dir:")
                for f in avail:
                    lines.append("    - " + f)
        except OSError:
            pass

    return valid, missing, lines
